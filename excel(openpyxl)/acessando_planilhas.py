import openpyxl     #'''importa openpyxl'''
from openpyxl.cell import get_column_letter, column_index_from_string
#'''importa funções para converter letras e números'''


wb = openpyxl.load_workbook('example_copy.xlsx')
sheet = wb.get_sheet_by_name('Plan1')		#obtém a planilha Plan1
anotherSheet = wb.active
conteudo = sheet['A3']


print ('Nome das planilhas: ', wb.get_sheet_names())
print ('Obtendo nome do objeto worksheet: ', wb.get_sheet_by_name('Problemas'))
print ('Qual o tipo da planilha: ', type(wb))
print ('Título da planilha: ', sheet.title)
print ('Planilha ativa: ', anotherSheet)
print ('Tipo da célula: ', sheet['A2'])
print ('Valor: ', sheet['A2'].value)
print ('Valor: ', conteudo.value)
print ('Row ' + str(conteudo.row) + ', Column ' + conteudo.column + ' é ' + conteudo.value)
print ('Cell ' + conteudo.coordinate + ' é ' + conteudo.value)

print ('', sheet.cell(row=1, column=2).value)
for i in range(1,10,1):
    print(i, sheet.cell(row=i, column=2).value)

print ('Tamanho máximo de linhas da planilha: ', str(sheet.max_row))
print ('Tamanho máximo de colunas da planilha: ', str(sheet.max_column))


#Convertendo letras e números
print ('Letra correspondente a coluna: ', get_column_letter(1))
print ('Letra correspondente a última coluna: ', get_column_letter(sheet.max_column))
print ('Letra correspondente a coluna: ', column_index_from_string('AA'))
print (tuple(sheet['A1':'B10']))

#Obtendo valores das células em uma área retangular
for linhas in sheet['A1':'B10']:
    for colunas in linhas:
        print (colunas.coordinate, colunas.value)
    print ('-----FIM-----')

print ('', sheet.columns[0])

for celula in sheet.columns[0]:
    print (celula.value)


# sheet = wb.create_sheet()		#Cria nova planilha
# sheet = wb.create_sheet(index=0, title='First Sheet')		#Cria nova planilha com posição definida, no caso 1ª planilha
sheet = wb.remove_sheet(wb.get_sheet_by_name('Plan3'))		#Remove a planilha Plan1
wb.save('example_copy.xlsx')		#Salva os dados em uma cópia do arquivo. Melhor prática para não danificar o arquivo 
									#original. Só será salvo as modificações se existir esse código.
