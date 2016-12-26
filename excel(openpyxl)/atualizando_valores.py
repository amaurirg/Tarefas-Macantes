#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.
#! python3
# Atualizando valores em planilhas xlsx

import openpyxl		#importa o módulo openpyxl para trabalhar com planilhas no formato xlsx

wb = openpyxl.load_workbook('produceSales.xlsx')	#lê o documento, no caso, produceSales.xlsx
sheet = wb.get_sheet_by_name('Sheet')	#informa para trabalhar na planilha especificada, no caso, Sheet

PRICE_UPDATES = {}	#Cria dicionário vazio.

resp = ''	#declara variável para resposta 's' ou 'n' valor em branco

while resp != 'n':	#enquanto a resposta for diferente de 'n' faz o loop
	resp = input('\nAtualizar produto? (s para sim ou n para sair): ')		#Mostra mensagem na tela, agurarda 's' ou 'n' e armazena na variável 'resp'
	try:	#junto com except ValueError para tratar erro de valor digitado
		if resp == 's':		#se a resposta for 's' executa os comandos os abaixo
			produto = input('Digite o produto: ')	#Mostra mensagem na tela, agurarda o nome de um produto e armazena na variável produto
			preco = input('Digite o valor: ')		#Mostra mensagem na tela, agurarda o valor de um produto e armazena na variável preco
			
			#Necessário esse if/else para verificar pois o float converte no estilo americano (3.50) e não no modo brasileiro (3,50)
			#Se tentar converter (3,50) para float gera erro pois o Python considera uma tupla com dois valores '3' e '50'
			#No caso de um usuário digitar um valor com vírgula, executa o 'if'. Caso contrário, executa o else. E se digitar, por exemplo,
			#'qualquer_coisa,mais_qualquer_coisa' gera um ValueError que é tratado com try/execpt pedindo que o usuário digite um valor válido
			if ',' in preco:	#se tiver ',' no valor digitado executa os comandos os abaixo
				valor = preco.replace(',', '.')		#troca ',' por '.'
				PRICE_UPDATES[produto] = float(valor)	#converte o valor digitado em float e armazena como chave o valor da variável 'produto' 
														#e como valor o valor da variável 'valor' no dicionário PRICE_UPDATES
			else:	#qualquer caractere digitado exceto quando digitado ',' será tratado aqui
				PRICE_UPDATES[produto] = float(preco)	#converte o valor digitado em float e armazena como chave o valor da variável 'produto' 
														#e como valor o valor da variável 'preco' no dicionário PRICE_UPDATES
		else:
			print('Digite s para alterar um produto e valor ou n para sair')	#caso o usuário não digite 's' ou 'n' retorna no começo do loop
																				#while para aguardar uma resposta válida
	except ValueError:		#no caso do usuário digitar um valor inválido, gera um ValueError que é tratado pedindo que ele digite um valor válido
			preco = input('Digite um valor válido (ex.: 2,00): ')	#aguarda um valor válido e armazena na variável 'preco'
			
			#Necessário esse if/else para informar que no caso do usuário digitar um valor inválido por duas vezes, não armazenará o produto e valor
			#e reiniciará o loop while perguntando se deseja atualizar algum produto
			if not ValueError:	#se não houver erro de ValueError, não faz nada e continua executando o programa
				pass 	#não faz nada
			else:	#se houver erro de ValueError, pede novo valor e caso persista, reinicia o loop while
				print('O valor desse produto não será alterado devido ao valor inválido informado.')	#mensagem para informar que devido o valor
				#inválido não será armazenado no dicionário e portanto não substituirá o valor do produto
print(PRICE_UPDATES)	#mostra o dicionário com os valores que serão alterados


for rowNum in range(2, sheet.max_row+1): #percorre o documento da 2ª linha até a última. Tem que colocar '+1' pois o range seria até a penúltima linha
    produceName = sheet.cell(row=rowNum, column=1).value 	#percorre as linhas da coluna 1 e armazena o valor da célula em 'produceName'
    if produceName in PRICE_UPDATES:	#se o produto for encontrado na coluna 1 executa o comando abaixo
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]		#altera o valor do produto na coluna 2

wb.save('updatedProduceSales.xlsx')		#comando necessário para salvar efetivamente as alterações na planilha